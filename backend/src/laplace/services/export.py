import csv
import io
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_xlsx_report(
    data: dict,
    diagnostics: dict | None,
    backtest: dict | None,
    forecast: dict | None,
    sections: list[str] | None = None,
    notes: str | None = None,
) -> bytes:
    """Generate XLSX report. sections controls which sheets are included."""
    all_sections = {"summary", "forecast", "backtest", "diagnostics", "raw_data", "notes"}
    include = set(sections) if sections else all_sections

    wb = Workbook()

    if "summary" in include:
        _write_summary_sheet(wb, data, diagnostics, backtest, forecast)
    else:
        # openpyxl always has a default sheet; rename it to a placeholder then remove after
        ws = wb.active
        ws.title = "_placeholder"

    if "forecast" in include:
        _write_forecast_sheet(wb, data, forecast)
    if "backtest" in include:
        _write_backtest_sheet(wb, backtest)
    if "diagnostics" in include:
        _write_diagnostics_sheet(wb, diagnostics)
    if "raw_data" in include:
        _write_raw_data_sheet(wb, data)
    if "notes" in include and notes:
        _write_notes_sheet(wb, notes)

    # Remove placeholder sheet if summary was skipped
    if "_placeholder" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["_placeholder"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_csv_report(
    data: dict,
    backtest: dict | None,
    forecast: dict | None,
) -> bytes:
    """Generate a flat CSV with the most important data for quick use."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header block
    writer.writerow(["# Laplace Export", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")])
    writer.writerow(["# Dataset", data.get("name", ""), "Frequency", data.get("frequency", ""), "Points", data.get("n_points", "")])
    writer.writerow([])

    # Backtest metrics block
    if backtest:
        writer.writerow(["# Backtest Aggregate Metrics"])
        writer.writerow(["Model", "sMAPE", "MAE", "RMSE", "MAPE", "MASE"])
        winner = backtest.get("winner", "")
        for name, m in backtest.get("aggregate_metrics", {}).items():
            star = "*" if name == winner else ""
            writer.writerow([
                f"{name}{star}",
                round(m.get("smape", 0), 4),
                round(m.get("mae", 0), 4),
                round(m.get("rmse", 0), 4),
                round(m.get("mape") or 0, 4),
                round(m.get("mase", 0), 4),
            ])
        writer.writerow([])

    # Forecast block
    if forecast and forecast.get("point_forecast"):
        writer.writerow(["# Forecast"])
        writer.writerow(["Step", "Point Forecast", "Lo 80%", "Hi 80%", "Lo 90%", "Hi 90%"])
        pf = forecast["point_forecast"]
        for i in range(len(pf)):
            writer.writerow([
                i + 1,
                round(pf[i], 4),
                round(forecast["lo_80"][i], 4),
                round(forecast["hi_80"][i], 4),
                round(forecast["lo_90"][i], 4),
                round(forecast["hi_90"][i], 4),
            ])
        writer.writerow([])

    # Raw data block
    dates = data.get("dates", [])
    values = data.get("values", [])
    if values:
        writer.writerow(["# Raw Data"])
        writer.writerow(["Date", "Value"])
        for i, v in enumerate(values):
            writer.writerow([dates[i] if i < len(dates) else "", round(v, 6)])

    return buf.getvalue().encode("utf-8")


def read_results_log() -> list[dict]:
    """Read all entries from the results log CSV."""
    if not RESULTS_LOG_PATH.exists():
        return []
    with open(RESULTS_LOG_PATH, newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def delete_log_entry(index: int) -> None:
    """Delete a single entry from the log by 0-based index."""
    entries = read_results_log()
    if index < 0 or index >= len(entries):
        raise IndexError(f"Log entry index {index} out of range (0–{len(entries) - 1})")
    entries.pop(index)
    _write_results_log(entries)


def clear_results_log() -> None:
    """Delete all entries (remove the log file)."""
    if RESULTS_LOG_PATH.exists():
        RESULTS_LOG_PATH.unlink()


def _write_results_log(entries: list[dict]) -> None:
    """Overwrite the log with a new list of entries."""
    with open(RESULTS_LOG_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=LOG_HEADERS)
        writer.writeheader()
        writer.writerows(entries)


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="E8E8ED", end_color="E8E8ED", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="DDDDDD"))
NUM_FMT = "#,##0.00"


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _write_summary_sheet(wb, data, diagnostics, backtest, forecast):
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ("Dataset", data.get("name", "")),
        ("Points", data.get("n_points", len(data.get("values", [])))),
        ("Frequency", data.get("frequency", "")),
        ("Generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
    ]

    if diagnostics:
        fc = diagnostics.get("forecastability", {})
        rows.append(("Forecastability Score", fc.get("total_score", "")))
        rows.append(("Interpretation", fc.get("interpretation", "")))
        rows.append(("", ""))

    if backtest:
        rows.append(("Backtest Winner", backtest.get("winner", "")))
        rows.append(("Selection Metric", backtest.get("selection_metric", "")))
        rows.append(("CV Folds", backtest.get("n_splits", "")))
        rows.append(("Horizon", backtest.get("horizon", "")))
        rows.append(("", ""))

        for name, m in backtest.get("aggregate_metrics", {}).items():
            rows.append((f"{name} sMAPE", m.get("smape", "")))

    if forecast:
        rows.append(("", ""))
        rows.append(("Forecast Model", forecast.get("model_name", "")))
        rows.append(("Forecast Horizon", forecast.get("horizon", "")))

    for r_idx, (label, value) in enumerate(rows, 1):
        ws.cell(row=r_idx, column=1, value=label).font = Font(bold=True) if label else Font()
        ws.cell(row=r_idx, column=2, value=value)

    _auto_width(ws)


def _write_forecast_sheet(wb, data, forecast):
    ws = wb.create_sheet("Forecast")

    if not forecast or not forecast.get("point_forecast"):
        ws.cell(row=1, column=1, value="No forecast generated")
        return

    headers = ["Step", "Point Forecast", "Lo 80%", "Hi 80%", "Lo 90%", "Hi 90%"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    n = len(forecast["point_forecast"])
    for i in range(n):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=forecast["point_forecast"][i]).number_format = NUM_FMT
        ws.cell(row=row, column=3, value=forecast["lo_80"][i]).number_format = NUM_FMT
        ws.cell(row=row, column=4, value=forecast["hi_80"][i]).number_format = NUM_FMT
        ws.cell(row=row, column=5, value=forecast["lo_90"][i]).number_format = NUM_FMT
        ws.cell(row=row, column=6, value=forecast["hi_90"][i]).number_format = NUM_FMT

    _auto_width(ws)


def _write_backtest_sheet(wb, backtest):
    ws = wb.create_sheet("Backtest Metrics")

    if not backtest:
        ws.cell(row=1, column=1, value="No backtest results")
        return

    headers = ["Model", "sMAPE (%)", "MAE", "RMSE", "MAPE (%)", "MASE"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    winner = backtest.get("winner", "")
    for r_idx, (name, m) in enumerate(backtest.get("aggregate_metrics", {}).items(), 2):
        ws.cell(row=r_idx, column=1, value=name)
        ws.cell(row=r_idx, column=2, value=m.get("smape")).number_format = NUM_FMT
        ws.cell(row=r_idx, column=3, value=m.get("mae")).number_format = NUM_FMT
        ws.cell(row=r_idx, column=4, value=m.get("rmse")).number_format = NUM_FMT
        mape = m.get("mape")
        ws.cell(row=r_idx, column=5, value=mape if mape is not None else "N/A")
        ws.cell(row=r_idx, column=6, value=m.get("mase")).number_format = "0.000"

        if name == winner:
            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).font = Font(bold=True, color="0066FF")

    _auto_width(ws)


def _write_diagnostics_sheet(wb, diagnostics):
    ws = wb.create_sheet("Diagnostics")

    if not diagnostics:
        ws.cell(row=1, column=1, value="No diagnostics data")
        return

    fc = diagnostics.get("forecastability", {})
    dims = fc.get("dimensions", [])

    headers = ["Dimension", "Score", "Weight", "Description"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for r_idx, d in enumerate(dims, 2):
        ws.cell(row=r_idx, column=1, value=d.get("name", ""))
        ws.cell(row=r_idx, column=2, value=d.get("score")).number_format = "0.0"
        ws.cell(row=r_idx, column=3, value=f"{d.get('weight', 0) * 100:.0f}%")
        ws.cell(row=r_idx, column=4, value=d.get("description", ""))

    row = len(dims) + 3
    ws.cell(row=row, column=1, value="Total Score").font = Font(bold=True)
    ws.cell(row=row, column=2, value=fc.get("total_score")).number_format = "0.0"

    _auto_width(ws)


def _write_raw_data_sheet(wb, data):
    ws = wb.create_sheet("Raw Data")

    headers = ["Date", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    dates = data.get("dates", [])
    values = data.get("values", [])
    for i in range(len(values)):
        row = i + 2
        ws.cell(row=row, column=1, value=dates[i] if i < len(dates) else "")
        ws.cell(row=row, column=2, value=values[i]).number_format = NUM_FMT

    _auto_width(ws)


def _write_notes_sheet(wb, notes: str):
    ws = wb.create_sheet("Analyst Notes")
    ws.cell(row=1, column=1, value="Analyst Notes").font = HEADER_FONT
    for i, line in enumerate(notes.splitlines(), 2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 80


RESULTS_LOG_PATH = Path(__file__).resolve().parents[3] / "results_log.csv"
LOG_HEADERS = [
    "timestamp",
    "dataset",
    "model",
    "smape",
    "mae",
    "rmse",
    "horizon",
    "forecastability_score",
    "n_observations",
]


def append_results_log(entry: dict) -> str:
    file_exists = RESULTS_LOG_PATH.exists()

    with open(RESULTS_LOG_PATH, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=LOG_HEADERS)
        if not file_exists:
            writer.writeheader()
        row = {
            "timestamp": datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S"),
            "dataset": entry.get("dataset", ""),
            "model": entry.get("model", ""),
            "smape": entry.get("smape", ""),
            "mae": entry.get("mae", ""),
            "rmse": entry.get("rmse", ""),
            "horizon": entry.get("horizon", ""),
            "forecastability_score": entry.get("forecastability_score", ""),
            "n_observations": entry.get("n_observations", ""),
        }
        writer.writerow(row)

    return str(RESULTS_LOG_PATH)
