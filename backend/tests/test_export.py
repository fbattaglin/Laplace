import csv

import pytest
from openpyxl import load_workbook

from laplace.services.export import (
    append_results_log,
    clear_results_log,
    delete_log_entry,
    generate_xlsx_report,
    read_results_log,
)


class TestXlsxReport:
    def test_generates_valid_workbook(self):
        data = {
            "name": "test_series",
            "dates": ["2024-01", "2024-02", "2024-03"],
            "values": [10.0, 20.0, 30.0],
            "frequency": "M",
            "n_points": 3,
        }
        xlsx_bytes = generate_xlsx_report(data, None, None, None)
        assert len(xlsx_bytes) > 0

        import io

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) == 5
        assert "Summary" in wb.sheetnames
        assert "Forecast" in wb.sheetnames
        assert "Backtest Metrics" in wb.sheetnames
        assert "Diagnostics" in wb.sheetnames
        assert "Raw Data" in wb.sheetnames

    def test_raw_data_sheet_correct(self):
        data = {
            "name": "test",
            "dates": ["2024-01", "2024-02"],
            "values": [100.0, 200.0],
            "frequency": "M",
        }
        import io

        wb = load_workbook(io.BytesIO(generate_xlsx_report(data, None, None, None)))
        ws = wb["Raw Data"]
        assert ws.cell(row=1, column=1).value == "Date"
        assert ws.cell(row=1, column=2).value == "Value"
        assert ws.cell(row=2, column=1).value == "2024-01"
        assert ws.cell(row=2, column=2).value == 100.0

    def test_backtest_sheet_with_data(self):
        data = {"name": "test", "dates": [], "values": [], "frequency": "M"}
        backtest = {
            "winner": "Chronos-2",
            "selection_metric": "smape",
            "n_splits": 3,
            "horizon": 12,
            "aggregate_metrics": {
                "Chronos-2": {"smape": 5.0, "mae": 10.0, "rmse": 12.0, "mape": 4.5, "mase": 0.5},
                "AutoETS": {"smape": 8.0, "mae": 15.0, "rmse": 18.0, "mape": 7.0, "mase": 0.8},
            },
        }
        import io

        wb = load_workbook(io.BytesIO(generate_xlsx_report(data, None, backtest, None)))
        ws = wb["Backtest Metrics"]
        assert ws.cell(row=2, column=1).value == "Chronos-2"
        assert ws.cell(row=2, column=2).value == 5.0

    def test_forecast_sheet_with_data(self):
        data = {"name": "test", "dates": [], "values": [], "frequency": "M"}
        forecast = {
            "model_name": "Chronos-2",
            "horizon": 3,
            "point_forecast": [100.0, 110.0, 120.0],
            "lo_80": [90.0, 100.0, 110.0],
            "hi_80": [110.0, 120.0, 130.0],
            "lo_90": [85.0, 95.0, 105.0],
            "hi_90": [115.0, 125.0, 135.0],
        }
        import io

        wb = load_workbook(io.BytesIO(generate_xlsx_report(data, None, None, forecast)))
        ws = wb["Forecast"]
        assert ws.cell(row=1, column=2).value == "Point Forecast"
        assert ws.cell(row=2, column=2).value == 100.0
        assert ws.cell(row=4, column=2).value == 120.0


class TestResultsLog:
    def test_creates_and_appends(self, tmp_path, monkeypatch):
        log_path = tmp_path / "results_log.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

        entry = {
            "dataset": "airline",
            "model": "Chronos-2",
            "smape": 5.0,
            "mae": 10.0,
            "rmse": 12.0,
            "horizon": 12,
            "forecastability_score": 77.0,
            "n_observations": 144,
        }
        append_results_log(entry)
        assert log_path.exists()

        with open(log_path) as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["dataset"] == "airline"
        assert rows[0]["model"] == "Chronos-2"

    def test_appends_multiple(self, tmp_path, monkeypatch):
        log_path = tmp_path / "results_log.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

        for i in range(3):
            append_results_log({"dataset": f"ds_{i}", "model": "test"})

        with open(log_path) as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 3

    def test_delete_entry_by_index(self, tmp_path, monkeypatch):
        log_path = tmp_path / "results_log.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

        for i in range(3):
            append_results_log({"dataset": f"ds_{i}", "model": "test"})

        delete_log_entry(1)  # delete middle entry (ds_1)

        remaining = read_results_log()
        assert len(remaining) == 2
        assert remaining[0]["dataset"] == "ds_0"
        assert remaining[1]["dataset"] == "ds_2"

    def test_delete_out_of_range_raises(self, tmp_path, monkeypatch):
        log_path = tmp_path / "results_log.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

        append_results_log({"dataset": "only_one", "model": "test"})

        with pytest.raises(IndexError):
            delete_log_entry(5)

    def test_clear_log_removes_file(self, tmp_path, monkeypatch):
        log_path = tmp_path / "results_log.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

        append_results_log({"dataset": "to_be_cleared", "model": "test"})
        assert log_path.exists()

        clear_results_log()
        assert not log_path.exists()
        assert read_results_log() == []

    def test_clear_nonexistent_log_no_error(self, tmp_path, monkeypatch):
        log_path = tmp_path / "does_not_exist.csv"
        monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)
        clear_results_log()  # should not raise


@pytest.mark.asyncio
async def test_export_xlsx_endpoint(client):
    response = await client.post(
        "/api/export/xlsx",
        json={
            "data": {
                "name": "test",
                "dates": ["2024-01", "2024-02"],
                "values": [10, 20],
                "frequency": "M",
            }
        },
    )
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert len(response.content) > 100


@pytest.mark.asyncio
async def test_delete_log_entry_endpoint(client, tmp_path, monkeypatch):
    log_path = tmp_path / "results_log.csv"
    monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

    log_entry = {
        "dataset": "airline", "model": "Chronos-2", "smape": 5.0,
        "mae": 10.0, "rmse": 12.0, "horizon": 12,
        "forecastability_score": 77.0, "n_observations": 144, "frequency": "M",
    }
    # Save two entries
    await client.post("/api/export/log", json=log_entry)
    await client.post("/api/export/log", json={**log_entry, "dataset": "energy_demand"})

    # Delete the first entry (index 0)
    resp = await client.delete("/api/export/log/0")
    assert resp.status_code == 200

    remaining = read_results_log()
    assert len(remaining) == 1
    assert remaining[0]["dataset"] == "energy_demand"


@pytest.mark.asyncio
async def test_delete_log_entry_out_of_range(client, tmp_path, monkeypatch):
    log_path = tmp_path / "results_log.csv"
    monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

    resp = await client.delete("/api/export/log/99")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_clear_log_endpoint(client, tmp_path, monkeypatch):
    log_path = tmp_path / "results_log.csv"
    monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

    log_entry = {
        "dataset": "airline", "model": "Chronos-2", "smape": 5.0,
        "mae": 10.0, "rmse": 12.0, "horizon": 12,
        "forecastability_score": 77.0, "n_observations": 144, "frequency": "M",
    }
    await client.post("/api/export/log", json=log_entry)

    resp = await client.delete("/api/export/log")
    assert resp.status_code == 200

    # Log should be empty now
    resp = await client.get("/api/export/log")
    assert resp.json()["entries"] == []


@pytest.mark.asyncio
async def test_export_log_endpoint(client, tmp_path, monkeypatch):
    log_path = tmp_path / "results_log.csv"
    monkeypatch.setattr("laplace.services.export.RESULTS_LOG_PATH", log_path)

    response = await client.post(
        "/api/export/log",
        json={
            "dataset": "airline",
            "model": "Chronos-2",
            "smape": 5.0,
            "mae": 10.0,
            "rmse": 12.0,
            "horizon": 12,
            "forecastability_score": 77.0,
            "n_observations": 144,
            "frequency": "M",
        },
    )
    assert response.status_code == 200
    assert response.json()["status"] == "ok"
